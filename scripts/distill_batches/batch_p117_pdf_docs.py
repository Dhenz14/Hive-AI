"""Document generation — PDF creation, spreadsheet processing, and report generation."""

PAIRS = [
    (
        "python/pdf-generation",
        "Show PDF generation patterns: creating PDFs with reportlab, HTML-to-PDF, and invoice/report templates.",
        '''PDF generation patterns:

```python
# --- reportlab: programmatic PDF creation ---

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch, mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, Image, PageBreak,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from io import BytesIO
from datetime import date


def generate_invoice(invoice_data: dict) -> bytes:
    """Generate a PDF invoice."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        rightMargin=0.75 * inch,
        leftMargin=0.75 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Heading1"],
        fontSize=24, textColor=colors.HexColor("#1e40af"),
    )
    heading_style = ParagraphStyle(
        "SubHeading", parent=styles["Heading2"],
        fontSize=14, spaceAfter=6,
    )

    elements = []

    # Header
    elements.append(Paragraph("INVOICE", title_style))
    elements.append(Spacer(1, 12))

    # Invoice details (2-column layout)
    info_data = [
        ["Invoice #:", invoice_data["number"]],
        ["Date:", invoice_data["date"]],
        ["Due Date:", invoice_data["due_date"]],
    ]
    info_table = Table(info_data, colWidths=[100, 200])
    info_table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 24))

    # Bill to
    elements.append(Paragraph("Bill To:", heading_style))
    elements.append(Paragraph(invoice_data["customer_name"], styles["Normal"]))
    elements.append(Paragraph(invoice_data["customer_address"], styles["Normal"]))
    elements.append(Spacer(1, 24))

    # Line items table
    header = ["Description", "Qty", "Unit Price", "Total"]
    table_data = [header]

    for item in invoice_data["items"]:
        total = item["quantity"] * item["unit_price"]
        table_data.append([
            item["description"],
            str(item["quantity"]),
            f"${item['unit_price']:.2f}",
            f"${total:.2f}",
        ])

    # Subtotal, tax, total
    subtotal = sum(i["quantity"] * i["unit_price"] for i in invoice_data["items"])
    tax = subtotal * invoice_data.get("tax_rate", 0)
    total = subtotal + tax

    table_data.append(["", "", "Subtotal:", f"${subtotal:.2f}"])
    table_data.append(["", "", f"Tax ({invoice_data.get('tax_rate', 0):.0%}):", f"${tax:.2f}"])
    table_data.append(["", "", "Total:", f"${total:.2f}"])

    items_table = Table(table_data, colWidths=[250, 60, 100, 100])
    items_table.setStyle(TableStyle([
        # Header row
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        # Body
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        # Alternating row colors
        ("ROWBACKGROUNDS", (0, 1), (-1, -4), [colors.white, colors.HexColor("#f8fafc")]),
        # Grid
        ("GRID", (0, 0), (-1, -4), 0.5, colors.HexColor("#e2e8f0")),
        # Totals section
        ("FONTNAME", (2, -3), (2, -1), "Helvetica-Bold"),
        ("FONTNAME", (3, -1), (3, -1), "Helvetica-Bold"),
        ("FONTSIZE", (2, -1), (3, -1), 12),
        ("LINEABOVE", (2, -3), (-1, -3), 1, colors.HexColor("#1e40af")),
        # Align amounts right
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
    ]))

    elements.append(items_table)
    elements.append(Spacer(1, 48))

    # Footer
    elements.append(Paragraph(
        "Thank you for your business!",
        ParagraphStyle("Footer", parent=styles["Normal"],
                       alignment=TA_CENTER, textColor=colors.gray),
    ))

    doc.build(elements)
    return buffer.getvalue()


# --- HTML to PDF (weasyprint) ---

def html_to_pdf(html_content: str) -> bytes:
    """Convert HTML/CSS to PDF using WeasyPrint."""
    from weasyprint import HTML

    return HTML(string=html_content).write_pdf()


# Usage with Jinja2 templates:
def render_report_pdf(template_name: str, data: dict) -> bytes:
    from jinja2 import Environment, FileSystemLoader
    from weasyprint import HTML

    env = Environment(loader=FileSystemLoader("templates"))
    template = env.get_template(template_name)
    html = template.render(**data)

    return HTML(string=html, base_url="templates/").write_pdf()


# --- FastAPI endpoint ---

from fastapi import FastAPI
from fastapi.responses import Response

app = FastAPI()

@app.get("/invoices/{invoice_id}/pdf")
async def download_invoice(invoice_id: str):
    invoice = await get_invoice(invoice_id)
    pdf_bytes = generate_invoice(invoice)

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=invoice-{invoice_id}.pdf",
        },
    )
```

PDF generation patterns:
1. **reportlab `Platypus`** — flow-based layout with Table, Paragraph, Spacer
2. **`TableStyle`** — cell-level formatting: colors, fonts, alignment, grid lines
3. **WeasyPrint** — HTML/CSS to PDF (use Jinja2 for complex layouts)
4. **`BytesIO` buffer** — generate in memory, return as HTTP response
5. **Content-Disposition** — `attachment` triggers download, `inline` shows in browser'''
    ),
    (
        "python/spreadsheet-processing",
        "Show spreadsheet processing patterns: reading/writing Excel with openpyxl, CSV processing, and data export.",
        '''Spreadsheet processing patterns:

```python
import csv
import json
from io import BytesIO, StringIO
from pathlib import Path
from datetime import datetime
from typing import Iterator

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter


# --- Excel: create formatted spreadsheet ---

def create_report_xlsx(data: list[dict], title: str = "Report") -> bytes:
    """Create a formatted Excel report."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title

    if not data:
        return save_workbook(wb)

    headers = list(data[0].keys())

    # Styles
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write data
    for row_idx, record in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = record.get(header)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

            # Format numbers
            if isinstance(value, float):
                cell.number_format = '#,##0.00'
            elif isinstance(value, datetime):
                cell.number_format = 'YYYY-MM-DD HH:MM'

    # Auto-fit column widths
    for col in range(1, len(headers) + 1):
        max_length = max(
            len(str(ws.cell(row=row, column=col).value or ""))
            for row in range(1, len(data) + 2)
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max_length + 4, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"

    return save_workbook(wb)


def save_workbook(wb: openpyxl.Workbook) -> bytes:
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --- Excel: read and process ---

def read_xlsx(path: Path, sheet_name: str | None = None) -> list[dict]:
    """Read Excel file into list of dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip().lower().replace(" ", "_") for h in next(rows)]

    records = []
    for row in rows:
        record = dict(zip(headers, row))
        # Skip empty rows
        if any(v is not None for v in record.values()):
            records.append(record)

    wb.close()
    return records


# --- CSV: streaming processing ---

def process_csv_stream(
    path: Path,
    transform_fn,
    output_path: Path,
    batch_size: int = 1000,
) -> dict:
    """Process large CSV in streaming batches."""
    stats = {"read": 0, "written": 0, "skipped": 0}

    with open(path, newline="", encoding="utf-8") as infile, \
         open(output_path, "w", newline="", encoding="utf-8") as outfile:

        reader = csv.DictReader(infile)
        writer = None

        for row in reader:
            stats["read"] += 1
            result = transform_fn(row)

            if result is None:
                stats["skipped"] += 1
                continue

            if writer is None:
                writer = csv.DictWriter(outfile, fieldnames=result.keys())
                writer.writeheader()

            writer.writerow(result)
            stats["written"] += 1

    return stats


# --- Multi-sheet Excel export ---

def create_multi_sheet_report(sheets: dict[str, list[dict]]) -> bytes:
    """Create Excel workbook with multiple sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for sheet_name, data in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])  # Max 31 chars
        if not data:
            continue

        headers = list(data[0].keys())
        ws.append([h.replace("_", " ").title() for h in headers])

        for record in data:
            ws.append([record.get(h) for h in headers])

    return save_workbook(wb)

# Usage:
# xlsx_bytes = create_multi_sheet_report({
#     "Users": users_data,
#     "Orders": orders_data,
#     "Revenue": revenue_data,
# })
```

Spreadsheet patterns:
1. **openpyxl styles** — Font, PatternFill, Border for professional formatting
2. **`freeze_panes`** — freeze header row for scrolling large datasets
3. **Auto-fit columns** — calculate width from max cell content length
4. **`read_only=True`** — memory-efficient reading for large files
5. **Streaming CSV** — `DictReader`/`DictWriter` process row-by-row without loading all'''
    ),
]
